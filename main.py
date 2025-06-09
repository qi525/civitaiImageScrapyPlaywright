# -*- coding: utf-8 -*-
#
# 功能模块和实现手法整理
#
# 本脚本旨在通过自动化浏览器（Playwright）抓取 Civitai.com 图片搜索结果页面的信息，
# 包括每个搜索结果的图片缩略图（并下载到本地）、其关联的原始图片详情页链接，
# 以及5个表情符号（点赞、爱心、笑哭、伤心、打赏）及其对应的数值。
# 抓取到的数据将与抓取时间戳、搜索关键词（本次固定为`tags=4`）和搜索URL一同被整理并导出到Excel文件中。
# 每个图片将保存到其专属的本地文件夹中，并在Excel中提供本地图片文件的绝对路径和可点击的超链接。
# 脚本会保留原有的日志系统，自动打开日志和结果文件，并支持注入Cookies。
#
# --- 主要功能模块 ---
#
# 1. 配置管理 (Configuration):
#    - PROXY: 定义HTTP代理地址，用于Playwright浏览器和aiohttp库的网络请求，规避地理限制或提高访问稳定性。
#    - LOG_DIR, IMAGE_DIR_BASE, RESULTS_DIR: 定义日志文件、图片存储和Excel结果文件的输出目录。
#    - KEYWORD_TARGET_FILE: 存储待搜索的关键词列表的文件 (在此脚本中，Civitai URL固定，此文件用于演示，实际不用于动态搜索).
#    - 目录创建: 脚本启动时自动创建所需的日志、图片和结果目录，确保文件能正确保存。
#    - 文件命名: 使用时间戳为日志和Excel文件生成唯一名称。
#
# 2. 日志系统 (Logging Setup):
#    - 使用Python内置的 `logging` 模块，配置日志记录器 `civitai_scraper`。
#    - 日志级别设置为 `INFO`，记录重要操作和信息。
#    - 同时配置文件处理器 (`FileHandler`) 和控制台处理器 (`StreamHandler`)，实现日志同时输出到文件和控制台。
#    - 日志格式化: 定义统一的日志输出格式。
#    - 目的: 方便跟踪脚本运行状态、调试问题和记录抓取过程中的事件。
#    - 自动打开日志: 脚本运行结束后，会自动尝试打开本次运行生成的日志文件。
#
# 3. 关键词读取 (Keyword Reading Helper Function - `read_keywords_from_file`):
#    - 目的: 从 `keywordTarget.txt` 文件中读取待搜索的关键词列表。
#    - 实现手法: 按行读取文件内容，每行视为一个关键词，并进行基本校验。
#    - 优点: 提高脚本的灵活性和可配置性。 (在此脚本中，由于Civitai URL固定，此函数仅为保留结构)
#
# 4. 浏览器自动化与数据抓取 (Browser Automation & Data Scraping - `performCivitaiImageScrape`):
#    - 核心模块，负责实际的网页交互和数据提取。
#    - 异步操作: 使用 `asyncio` 和 `playwright.async_api` 实现异步并发抓取。
#    - 浏览器上下文管理: 为每个抓取任务创建独立的浏览器上下文。
#    - 注入 Cookies: 尝试从 `cookies.json` 文件中读取并注入 cookies，以维持会话状态。
#    - 页面导航: 导航到Civitai图片搜索URL（`https://civitai.com/images?tags=4`）。
#    - 动态内容加载 (滚动): 实现瀑布流页面的动态加载，通过滚动页面触发新内容。
#    - 数据提取 (使用BeautifulSoup辅助):
#      - 定位大的图片结果元素（基于提供的CSS选择器）。
#      - 提取缩略图 (`img` 标签的 `src`)。
#      - 提取原始图片详情页链接 (`img` 标签的父级 `a` 标签的 `href`)。
#      - 提取5个表情及其数值 (使用正则表达式解析)。
#    - 图片下载: 使用 `aiohttp` 异步下载缩略图，并保存到本地文件夹中。
#      - 文件命名: 使用时间戳和图片URL的SHA256哈希值，确保唯一性。
#    - 数据存储: 将抓取到的数据存储到全局列表 `all_search_results_data` 中。
#    - 线程安全: 使用 `asyncio.Lock` 保护全局共享数据结构。
#    - **去重机制**: 使用 `download_history` 字典和图片内容的 MD5 哈希值，避免重复下载和存储相同的图片数据。
#    - **关键词提取**: 尝试从页面输入框中提取关键词。
#
# 5. 主执行逻辑 (Main Execution Logic - `main`):
#    - 启动Playwright浏览器，并设置固定的窗口分辨率。
#    - (可选) 调用 `read_keywords_from_file` 获取所有目标关键词 (在此脚本中，仅为结构保留)。
#    - 为每个关键词（或本次固定的Civitai URL）创建 `performCivitaiImageScrape` 任务，并并发执行。
#    - 关闭浏览器实例。
#
# 6. Excel数据导出 (Excel Export Logic):
#    - 使用 `openpyxl` 库创建新的Excel工作簿，并创建名为 "Civitai图片结果" 的工作表。
#    - 字段: "抓取时间", "搜索URL", "缩略图URL", "本地缩略图路径", "本地缩略图超链接", "原始图片详情页链接", "点赞数", "爱心数", "笑哭数", "伤心数", "打赏数", **"关键词"**。
#    - 超链接处理: 为搜索URL、原始图片详情页链接和**本地缩略图超链接**添加可点击的超链接，**并修改本地缩略图超链接的显示文本**。
#    - 列宽自适应: 根据列内容的最大长度自动调整列宽。
#
# 7. 脚本入口点 (Script Entry Point - `if __name__ == '__main__':`):
#    - 运行主异步函数。
#    - 错误处理: 捕获用户中断和所有未处理的异常，记录详细错误信息。
#    - 清理和自动打开文件 (`finally` 块): 关闭所有日志处理器，并尝试自动打开生成的日志文件和Excel结果文件。
#    - 兼容多操作系统: 自动判断操作系统并使用相应的命令打开文件。
#
# --- 技术栈 ---
# - Python 3.x
# - Playwright (异步浏览器自动化库)
# - BeautifulSoup4 (HTML解析库)
# - aiohttp (异步HTTP客户端，用于图片下载)
# - openpyxl (Excel文件读写库)
# - asyncio (Python异步编程框架)
# - logging (Python内置日志模块)
# - os, json, datetime, traceback, subprocess, hashlib, re (Python标准库)
# - aiofiles (异步文件操作，用于保存图片)
#
# --- 运行环境要求 ---
# - Python环境已安装。
# - 确保已安装所需的Python库: `pip install playwright beautifulsoup4 openpyxl aiohttp aiofiles`
# - 运行 `playwright install` 安装浏览器驱动。
# - 需要一个 `keywordTarget.txt` 文件，其中包含要搜索的关键词，一行一个。 (在此脚本中，此文件不再是必需的，因为URL固定)
# - **需要一个 `cookies.json` 文件，其中包含有效的 JSON 格式的 cookies 数据 (如果需要)。**
# - 可选配置代理 (`PROXY` 变量)。
#
import os
import json
import asyncio
import aiohttp
from datetime import datetime
import traceback
import logging
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink # Import Hyperlink for more control
import hashlib
import aiofiles
import base64
import re

# --- FIX FOR NameError: name 'playwright' is not defined ---
import playwright.async_api as playwright_api
from playwright.async_api import async_playwright, expect
# --- END FIX ---

from bs4 import BeautifulSoup


# --- Configuration ---
PROXY = "http://127.0.0.1:10808" # Your proxy address
LOG_DIR = "logs" # Directory for logs
IMAGE_DIR_BASE = "images_civitai" # Base directory for images - Changed for Civitai
RESULTS_DIR = "results_civitai" # Directory for Excel results - Changed for Civitai
KEYWORD_TARGET_FILE = "urlTarget.txt" # File to store target URLs for Civitai
DOWNLOAD_HISTORY_FILE = "download_history_civitai.json" # New: File to store download history for deduplication - Changed for Civitai

# Generate a timestamp for unique filenames - Moved to the top of the script for global access
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
log_filename = os.path.join(LOG_DIR, f"civitai_scraper_log_{timestamp}.txt") # Changed log filename
excel_filename = os.path.join(RESULTS_DIR, f"civitai_image_results_{timestamp}.xlsx") # Excel filename with timestamp


# --- Logging Setup ---
logger = logging.getLogger('civitai_scraper') # Changed logger name
logger.setLevel(logging.INFO)

# Create necessary directories if they don't exist
if not os.path.exists(IMAGE_DIR_BASE):
    os.makedirs(IMAGE_DIR_BASE) # Use makedirs to create intermediate directories if they don't exist
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
if not os.path.exists(RESULTS_DIR):
    os.makedirs(RESULTS_DIR)

# Create file handler for logging to a file
file_handler = logging.FileHandler(log_filename, encoding='utf-8')
file_handler.setLevel(logging.INFO)

# Create console handler for logging to console
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# Define log message format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# Add handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)
# --- End Logging Setup ---

# Global list to store data for Excel export
all_search_results_data = []
# Lock for thread-safe access to all_search_results_data when multiple async tasks are running
data_lock = asyncio.Lock()

# Global dictionary to store download history: {md5_hash_of_image_content: local_file_path}
download_history = {}


# --- Helper function to calculate MD5 hash of bytes ---
def calculate_md5(data_bytes):
    """Calculates the MD5 hash of a given byte string."""
    return hashlib.md5(data_bytes).hexdigest()

# --- Helper functions to load/save download history ---
def load_download_history(filepath):
    """Loads the download history from a JSON file."""
    global download_history
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                download_history = json.load(f)
            logger.info(f"Loaded download history from {filepath}")
        except json.JSONDecodeError:
            logger.warning(f"Warning: '{filepath}' contains invalid JSON. Starting with empty history.")
            download_history = {}
        except Exception as e:
            logger.error(f"Error loading download history from '{filepath}': {e}\n{traceback.format_exc()}")
            download_history = {}
    else:
        logger.info(f"Download history file '{filepath}' not found. Starting with empty history.")
        download_history = {}

def save_download_history(filepath):
    """Saves the current download history to a JSON file."""
    global download_history
    try:
        dir_name = os.path.dirname(filepath)
        if dir_name:  # 只有目录非空才创建
            os.makedirs(dir_name, exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(download_history, f, indent=4)
        logger.info(f"Saved download history to {filepath}")
    except Exception as e:
        logger.error(f"Error saving download history to '{filepath}': {e}\n{traceback.format_exc()}")

# --- New helper function to read URLs from file ---
# Renamed from read_keywords_from_file as we are now reading URLs
def read_urls_from_file(filepath):
    """Reads URLs from a text file, one URL per line."""
    urls = []
    if not os.path.exists(filepath):
        logger.error(f"Error: Target URL file '{filepath}' not found. Please create it with one URL per line. For example, add 'https://civitai.com/images?tags=4' to '{filepath}'.")
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                url = line.strip()
                if url and url.startswith("http"): # Basic validation for non-empty and http(s) URL
                    urls.append(url)
        if not urls:
            logger.warning(f"Warning: Target URL file '{filepath}' is empty or contains no valid URLs.")
        return urls
    except Exception as e:
        logger.error(f"Error reading URLs from '{filepath}': {e}\n{traceback.format_exc()}")
        return []
# --- End new helper function ---


async def process_image_data(image_url, base_folder_path):
    """
    Processes an image URL (external or data:image), handles deduplication using MD5,
    downloads if necessary, and returns the local file path and the content MD5 hash.
    If the image (by content MD5) is already in history, it returns the existing path.
    """
    local_filename = None
    image_content_md5 = None
    image_bytes = None

    if not image_url:
        logger.warning("Empty image URL skipped.")
        return None, None

    # Civitai images are typically external, so we'll focus on http/https
    if image_url.startswith('http'):
        url_without_query = image_url.split('?')[0]
        file_extension = url_without_query.split('.')[-1].lower()
        if not file_extension or len(file_extension) > 5 or not file_extension.isalpha():
            file_extension = 'jpg' # Default to jpg if no clear extension

        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(image_url, proxy=PROXY if PROXY else None, timeout=30.0) as response:
                    response.raise_for_status()
                    image_bytes = await response.read()
                    image_content_md5 = calculate_md5(image_bytes)

                    # Deduplication check
                    if image_content_md5 in download_history:
                        existing_path = download_history[image_content_md5]
                        logger.info(f"Downloaded image content (MD5: {image_content_md5}) already exists at: {existing_path}. Skipping save and using existing path.")
                        return existing_path, image_content_md5

                    local_filename = os.path.join(base_folder_path, f"{image_content_md5}.{file_extension}")

                    os.makedirs(os.path.dirname(local_filename), exist_ok=True)

                    async with aiofiles.open(local_filename, 'wb') as f:
                        await f.write(image_bytes)
                    logger.info(f"Image downloaded and saved: {local_filename}")

                    download_history[image_content_md5] = local_filename
                    return local_filename, image_content_md5

        except aiohttp.ClientResponseError as e:
            logger.error(f"HTTP error {e.status} downloading image {image_url}: {e.message}")
        except aiohttp.ClientConnectorError as e:
            logger.error(f"Network error (ClientConnectorError) downloading image {image_url}: {e}")
        except asyncio.TimeoutError:
            logger.error(f"Timeout error downloading image {image_url}")
        except Exception as e:
            logger.error(f"Unexpected error downloading image {image_url}: {e}\n{traceback.format_exc()}")
    else:
        logger.warning(f"Unsupported image URL format (not http/https): {image_url[:100]}...")

    return None, None


async def performCivitaiImageScrape(context, target_url):
    async_name = asyncio.current_task().get_name()

    # --- Load and add cookies to browser context ---
    try:
        if os.path.exists("cookies.json"):
            with open("cookies.json", "r", encoding="utf-8") as f:
                cookies = json.load(f)
                for cookie in cookies:
                    cookie_same_site = {'strict': 'Strict', 'Lax': 'Lax', 'none': 'None'}.get(cookie.get('sameSite'), None)
                    if cookie_same_site in ['Strict', 'Lax', 'None']:
                        cookie['sameSite'] = cookie_same_site
                    else:
                        if 'sameSite' in cookie:
                            del cookie['sameSite']
                await context.add_cookies(cookies)
            logger.info(f"{async_name} -> Cookies loaded and added to context.")
        else:
            logger.warning(f"{async_name} -> Warning: cookies.json not found. Proceeding without cookies. Please ensure 'cookies.json' exists if needed.")
    except json.JSONDecodeError:
        logger.error(f"{async_name} -> Error: Invalid JSON in cookies.json. Please check the file format. Full traceback:\n{traceback.format_exc()}")
    except Exception as e:
        logger.error(f"{async_name} -> Unexpected error loading cookies: {e}\n{traceback.format_exc()}")
    # --- End cookie loading ---

    page = await context.new_page()

    try:
        logger.info(f"{async_name} -> Navigating to {target_url}")
        await page.goto(target_url, timeout=60000, wait_until="domcontentloaded")
        logger.info(f"{async_name} -> Successfully navigated to {target_url}")
    except playwright_api.TimeoutError:
        logger.error(f"{async_name} -> Error: Page.goto timed out for {target_url} after 60 seconds. Check network or proxy.")
        await page.close()
        return
    except Exception as e:
        logger.error(f"{async_name} -> An unexpected error occurred during navigation to {target_url}: {e}")
        await page.close()
        return

    # Create a general Civitai image folder, as keywords are not used for subfolders here
    civitai_image_folder_path = os.path.join(IMAGE_DIR_BASE, "downloaded_images")
    if not os.path.exists(civitai_image_folder_path):
        os.makedirs(civitai_image_folder_path)
        logger.info(f"{async_name} -> Created base image folder for Civitai: {civitai_image_folder_path}")

    scrape_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    processed_image_detail_urls = set() # Set to keep track of processed image detail page URLs to avoid redundant processing

    # --- Dynamic Scrolling and Scraping ---
    scroll_attempts = 0
    max_scroll_attempts = 30 # Increased max attempts for more thorough scraping
    last_known_image_count = 0
    no_new_images_count = 0 # Counter for how many times no new images were found after scroll

    # Selector for the main image grid container
    image_boxes_selector = '#main > div > div > div > main > div.z-10.m-0.flex-1.mantine-1avyp1d > div > div > div > div:nth-child(2) > div.mx-auto.flex.justify-center.gap-4'
    
    # Keyword input selector
    keyword_input_selector = '#mantine-r6rf' # User provided CSS selector

    # Extract keyword from the input field if it exists
    current_keyword = "N/A"
    try:
        keyword_input_element = page.locator(keyword_input_selector)
        if await keyword_input_element.is_visible():
            input_value = await keyword_input_element.get_attribute('value')
            if input_value:
                current_keyword = input_value
                logger.info(f"{async_name} -> Found keyword in input field: '{current_keyword}'")
    except Exception as e:
        logger.warning(f"{async_name} -> Could not find or extract keyword from input field '{keyword_input_selector}': {e}")


    while scroll_attempts < max_scroll_attempts:
        scroll_attempts += 1
        logger.info(f"{async_name} -> Scroll attempt {scroll_attempts}/{max_scroll_attempts}...")

        # Scroll to the bottom of the page
        await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
        logger.info(f"{async_name} -> Scrolled to bottom. Waiting for content to load...")
        await asyncio.sleep(2) # Wait for content to load (increase if needed)

        # Optional: Scroll up slightly to ensure lazy-loaded images come into view
        # This simulates "上下来回中等速度进行翻动"
        if scroll_attempts % 3 == 0:
            await page.evaluate('window.scrollTo(0, document.body.scrollHeight * 0.7)')
            logger.info(f"{async_name} -> Scrolled up slightly (70%). Waiting 1 second...")
            await asyncio.sleep(1)

        # --- Data Extraction after each scroll ---
        try:
            # We need to get the outerHTML of each individual image box, not the container
            # The individual image box selector is:
            individual_image_box_selector = f'{image_boxes_selector} > div' # Each direct child div within the main grid
            
            # Wait for some elements to be attached, ensuring the DOM is somewhat ready
            await page.wait_for_selector(individual_image_box_selector, state='attached', timeout=15000)
            
            # Get all outer HTML of the current image boxes
            image_box_elements = await page.query_selector_all(individual_image_box_selector)
            
            current_image_count = len(image_box_elements)
            logger.info(f"{async_name} -> Found {current_image_count} image results on page after scroll attempt {scroll_attempts}.")

            newly_processed_this_scroll = 0

            for element in image_box_elements:
                html_content = await element.evaluate('node => node.outerHTML')
                soup = BeautifulSoup(html_content, "html.parser")

                result_data = {
                    "抓取时间": scrape_timestamp,
                    "搜索URL": target_url,
                    "缩略图URL": "",
                    "本地缩略图路径": "",
                    "本地缩略图超链接": "",
                    "原始图片详情页链接": "",
                    "点赞数": 0,
                    "爱心数": 0,
                    "笑哭数": 0,
                    "伤心数": 0,
                    "打赏数": 0,
                    "关键词": current_keyword # Add the extracted keyword
                }

                # Extract original image detail page link first for deduplication
                thumbnail_img_element = soup.select_one('img')
                original_page_link_element = thumbnail_img_element.find_parent('a') if thumbnail_img_element else None
                original_page_url = original_page_link_element.get('href') if original_page_link_element else ""
                if original_page_url and not original_page_url.startswith('http'):
                    original_page_url = f"https://civitai.com{original_page_url}"
                result_data["原始图片详情页链接"] = original_page_url

                # Deduplication check for already processed image *detail pages*
                if original_page_url and original_page_url in processed_image_detail_urls:
                    # logger.debug(f"{async_name} -> Skipping already processed image detail URL: {original_page_url}")
                    continue # Skip this entry if its detail page URL was already processed
                
                # Add to processed set ONLY IF we are going to process it
                if original_page_url:
                    processed_image_detail_urls.add(original_page_url)
                    newly_processed_this_scroll += 1


                # Now extract other details
                thumbnail_url = thumbnail_img_element.get('src') if thumbnail_img_element else ""
                result_data["缩略图URL"] = thumbnail_url
                
                # Process thumbnail image (download or use from history)
                if thumbnail_url:
                    local_image_path, image_md5 = await process_image_data(thumbnail_url, civitai_image_folder_path)
                    if local_image_path:
                        result_data["本地缩略图路径"] = os.path.abspath(local_image_path)
                        if os.name == 'nt':
                            formatted_local_path = result_data['本地缩略图路径'].replace('\\', '/')
                            result_data["本地缩略图超链接"] = f"file:///{formatted_local_path}"
                        else:
                            result_data["本地缩略图超链接"] = f"file://{result_data['本地缩略图路径']}"
                
                # Extract emoji data (5 buttons as per latest info)
                # Selector for the emoji container within the individual image box
                # It's 'div:nth-child(2) > div' relative to the 'current_image_box'
                emoji_container = soup.select_one('div:nth-child(2) > div')
                if emoji_container:
                    reaction_buttons = emoji_container.select('button')
                    for button in reaction_buttons:
                        count_span = button.find('span', class_=re.compile(r'mantine-h\w+'))
                        count = int(count_span.get_text(strip=True)) if count_span and count_span.get_text(strip=True).isdigit() else 0
                        
                        svg_element = button.find('svg')
                        
                        # --- Heuristics for SVG icons ---
                        # These are based on common Civitai icon patterns and might need refinement.
                        # The 'd' attribute of the <path> tag within the SVG is often unique for an icon.
                        
                        # Thumbs Up (点赞)
                        if svg_element and any(
                            'M13 14h-1.5a1 1 0 01-1-1V6.5a1 1 0 011-1H13a1 1 0 011 1V13a1 1 0 01-1 1z' in str(p) or # One example path
                            'M7 11.25V9.5h-1.5' in str(p) # Another common thumbs up path
                            for p in svg_element.find_all('path')
                        ):
                            result_data["点赞数"] = count
                        # Heart (爱心)
                        elif svg_element and any(
                            'M8 1.314C12.438-3.248 23.534 4.735 8 15-7.534 4.735 3.562-3.248 8 1.314z' in str(p) or # Standard heart path
                            'M16 8s-3.5 1.5-6.5 1.5S3 8 3 8l-.5-.5a4.7 4.7 0 010-6.7' in str(p) # Another possible heart path segment
                            for p in svg_element.find_all('path')
                        ):
                            result_data["爱心数"] = count
                        # Crying-Laughing (笑哭) - Often a specific face
                        elif svg_element and any(
                            'M8 15A7 7 0 108 1a7 7 0 000 14zm0 1A8 8 0 108 0a8 8 0 000 16z' in str(p) and # Face outline
                            'M7 8.5C7 8.224 7.224 8 7.5 8h1C8.776 8 9 8.224 9 8.5v.5H7V8.5z' in str(p) # Eye (part of crying face)
                            for p in svg_element.find_all('path')
                        ) or ('😂' in button.get_text()): # Fallback for unicode if present
                            result_data["笑哭数"] = count
                        # Sad (伤心) - Specific face
                        elif svg_element and any(
                            'M8 15A7 7 0 108 1a7 7 0 000 14zm0 1A8 8 0 108 0a8 8 0 000 16z' in str(p) and # Face outline
                            'M5 10.5h6a.5.5 0 010 1H5a.5.5 0 010-1z' in str(p) # Sad mouth line
                            for p in svg_element.find_all('path')
                        ) or ('😢' in button.get_text()): # Fallback for unicode if present
                            result_data["伤心数"] = count
                        # Yellow Lightning Bolt / Tip (打赏)
                        elif svg_element and any(
                            'M11.5 1h-8A1.5 1.5 0 002 2.5v11A1.5 1.5 0 003.5 15h8A1.5 1.5 0 0013 13.5v-11A1.5 1.5 0 0011.5 1zM12 2.5a.5.5 0 00-.5-.5h-8a.5.5 0 00-.5.5v11a.5.5 0 00.5.5h8a.5.5 0 00.5-.5v-11zM11 6a.5.5 0 00-.5.5v4a.5.5 0 001 0v-4A.5.5 0 0011 6zM5.5 6a.5.5 0 00-.5.5v4a.5.5 0 001 0v-4A.5.5 0 005.5 6z' in str(p) or # Common lightning bolt shape
                            'M6.5 1.5l-4 8h5l-1 5 7-9h-5l2-4z' in str(p) # Another lightning bolt shape
                            for p in svg_element.find_all('path')
                        ):
                            result_data["打赏数"] = count
                        else:
                            logger.debug(f"{async_name} -> Unknown reaction button found. Text: '{button.get_text(strip=True)}'. SVG excerpt: {str(svg_element)[:200] if svg_element else 'N/A'}")
                else:
                    logger.warning(f"{async_name} -> No emoji container (div:nth-child(2)>div) found for an image result.")

                # Only log and add if it's a newly processed item
                if original_page_url and original_page_url in processed_image_detail_urls: # Double check after adding
                    async with data_lock:
                        all_search_results_data.append(result_data)

            if newly_processed_this_scroll == 0 and current_image_count > 0:
                no_new_images_count += 1
                logger.info(f"{async_name} -> No new images processed this scroll. Consecutive no new images: {no_new_images_count}")
                if no_new_images_count >= 3: # If no new images for 3 consecutive scrolls, assume end
                    logger.info(f"{async_name} -> Reached end of content. Stopping scrolling.")
                    break
            else:
                no_new_images_count = 0 # Reset counter if new images were found

        except playwright_api.TimeoutError:
            logger.error(f"{async_name} -> Error: Image result elements not found on page after scroll within timeout. This might indicate no more content or a selector issue.")
            break # Stop scrolling if elements cannot be found
        except Exception as e:
            logger.error(f"{async_name} -> An unexpected error occurred during image results processing after scroll: {e}\n{traceback.format_exc()}")
            break # Stop scrolling on unexpected errors

    await page.close()
    logger.info(f"{async_name} -> Page closed for {target_url}.")


async def main():
    load_download_history(DOWNLOAD_HISTORY_FILE)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            proxy={"server": PROXY} if PROXY else None,
            timeout=60000
        )
        context = await browser.new_context(
            viewport={'width': 2560, 'height': 1440}
        )

        target_urls = read_urls_from_file(KEYWORD_TARGET_FILE)
        if not target_urls:
            logger.error(f"No valid URLs found in {KEYWORD_TARGET_FILE}. Please add URLs to scrape.")
            await browser.close()
            return

        tasks = [performCivitaiImageScrape(context, url) for url in target_urls]

        await asyncio.gather(*tasks)

        await browser.close()
        logger.info("Browser closed. Script finished scraping data.")

    # --- Excel Export Logic ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Civitai图片结果"

    # Define headers for the Civitai results sheet - Added "关键词"
    headers = ["抓取时间", "搜索URL", "缩略图URL", "本地缩略图路径", "本地缩略图超链接", "原始图片详情页链接", "点赞数", "爱心数", "笑哭数", "伤心数", "打赏数", "关键词"]
    ws.append(headers)

    hyperlink_font = Font(color="0000FF", underline="single")

    for row_data in all_search_results_data:
        row = []
        for header in headers:
            # For "本地缩略图超链接", we will set the cell value and hyperlink separately
            if header == "本地缩略图超链接":
                row.append("点击打开缩略图") # This will be the display text initially
            else:
                row.append(row_data.get(header, ""))
        ws.append(row)

        current_row_idx = ws.max_row

        # Apply hyperlink for "搜索URL"
        search_url = row_data.get("搜索URL")
        if search_url:
            cell_search_url = ws.cell(row=current_row_idx, column=headers.index("搜索URL") + 1)
            cell_search_url.value = search_url
            cell_search_url.hyperlink = search_url
            cell_search_url.font = hyperlink_font

        # Apply hyperlink for "缩略图URL"
        thumbnail_url = row_data.get("缩略图URL")
        if thumbnail_url:
            cell_thumbnail_url = ws.cell(row=current_row_idx, column=headers.index("缩略图URL") + 1)
            cell_thumbnail_url.value = thumbnail_url
            cell_thumbnail_url.hyperlink = thumbnail_url
            cell_thumbnail_url.font = hyperlink_font
            
        # Apply hyperlink for "本地缩略图超链接" with new display text and actual file link
        local_image_hyperlink_url = row_data.get("本地缩略图超链接") # This is the "file:///" URL
        if local_image_hyperlink_url:
            cell_local_image_hyperlink = ws.cell(row=current_row_idx, column=headers.index("本地缩略图超链接") + 1)
            cell_local_image_hyperlink.value = "点击打开缩略图" # Set display text
            # Create a Hyperlink object, linking to the file:// URL
            cell_local_image_hyperlink.hyperlink = Hyperlink(ref=local_image_hyperlink_url)
            cell_local_image_hyperlink.font = hyperlink_font


        # Apply hyperlink for "原始图片详情页链接"
        original_page_link = row_data.get("原始图片详情页链接")
        if original_page_link:
            cell_original_page_link = ws.cell(row=current_row_idx, column=headers.index("原始图片详情页链接") + 1)
            cell_original_page_link.value = original_page_link
            cell_original_page_link.hyperlink = original_page_link
            cell_original_page_link.font = hyperlink_font

    for col_idx, header in enumerate(headers):
        max_length = len(header)
        column_letter = get_column_letter(col_idx + 1)
        # Iterate over all cells in the column to find the true max_length, considering the new hyperlink text
        for r_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=r_idx, column=col_idx + 1).value
            if cell_value:
                cell_len = len(str(cell_value))
                if cell_len > max_length:
                    max_length = cell_len

        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > 100: # Cap column width
            adjusted_width = 100
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_filename)
    logger.info(f"Results saved to Excel: {excel_filename}")
    # --- END Excel Export Logic ---

    save_download_history(DOWNLOAD_HISTORY_FILE)


# --- 滚动相关函数建议放在这里 ---
async def civitai_basic_scroll(url="https://civitai.com/images?tags=4", scroll_times=30, sleep_sec=2):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={'width': 2560, 'height': 1440}
        )
        page = await context.new_page()
        await page.goto(url)
        for _ in range(scroll_times):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(sleep_sec)
        await browser.close()

async def civitai_scroll_all_elements(url="https://civitai.com/images?tags=4", max_rounds=50, sleep_sec=1):
    js_scroll_all = """
    document.querySelectorAll('*').forEach(function(el) {
      if (el.scrollHeight > el.clientHeight) el.scrollTop += 40;
    });
    """
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()
        await page.goto(url)
        last_height = 0
        for _ in range(max_rounds):
            await page.evaluate(js_scroll_all)
            await asyncio.sleep(sleep_sec)
            curr_height = await page.evaluate("document.body.scrollHeight")
            if curr_height == last_height:
                print("页面高度未增加，可能已滚到底部，停止滚动。")
                break
            last_height = curr_height
        await browser.close()
# --- END 滚动相关函数 ---


# --- New function to scroll and activate all elements on the page ---
async def civitai_scroll_all_elements(url="https://civitai.com/images?tags=4", max_rounds=50, sleep_sec=1):
    js_scroll_all = """
    document.querySelectorAll('*').forEach(function(el) {
      if (el.scrollHeight > el.clientHeight) el.scrollTop += 40;
    });
    """
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()
        await page.goto(url)
        last_height = 0
        for _ in range(max_rounds):
            await page.evaluate(js_scroll_all)
            await asyncio.sleep(sleep_sec)
            # 检查页面高度是否还在增加
            curr_height = await page.evaluate("document.body.scrollHeight")
            if curr_height == last_height:
                print("页面高度未增加，可能已滚到底部，停止滚动。")
                break
            last_height = curr_height
        await browser.close()

# 用法：asyncio.run(civitai_scroll_all_elements())
# --- END new function ---


async def inject_and_click_scroll_btn(url="https://civitai.com/images?tags=4", max_retry=10, retry_interval=1):
    with open("控制台注入版.js", "r", encoding="utf-8") as f:
        content_js_code = f.read()
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()
        page.on("console", lambda msg: print("PAGE LOG:", msg.text))
        await page.goto(url)
        await page.wait_for_load_state("networkidle")

        # 1. 先用 evaluate 注入
        for i in range(max_retry):
            print(f"[evaluate] 注入尝试 {i+1}/{max_retry} ...")
            await page.evaluate(content_js_code)
            #await page.evaluate("createMenu()")
            try:
                await page.wait_for_selector("#ultra-scroll-menu-btn", timeout=1000)
                print("[evaluate] 按钮已成功注入！")
                break
            except Exception:
                print("[evaluate] 按钮未出现，继续尝试注入...")
                await asyncio.sleep(retry_interval)
        else:
            print("[evaluate] 多次注入后仍未检测到按钮，尝试 add_init_script...")

            # 2. 用 add_init_script 注入
            await page.add_init_script(content_js_code)
            await page.reload()
            try:
                await page.wait_for_selector("#ultra-scroll-menu-btn", timeout=5000)
                print("[add_init_script] 按钮已成功注入！")
            except Exception:
                print("[add_init_script] 按钮未出现，尝试 expose_function...")

                # 3. 用 expose_function 方式（不推荐，但可演示）
                # 这里只能演示 expose_function 的用法，实际插入按钮还是得用 evaluate
                # expose_function 适合页面主动调用 Python，不适合直接插入按钮

                print("[expose_function] 暂无适用场景，建议用 evaluate 或 add_init_script。")
                await browser.close()
                return

        # 按钮出现后自动点击
        await page.click("#ultra-scroll-menu-btn")
        await asyncio.sleep(10)
        await page.click("#ultra-scroll-menu-btn")
        await asyncio.sleep(2)
        await browser.close()

# 用法
# asyncio.run(inject_and_click_scroll_btn())


if __name__ == '__main__':
    try:
        # 只测试按钮注入
        asyncio.run(inject_and_click_scroll_btn())
    except KeyboardInterrupt:
        logger.info("Script interrupted by user.")
    except Exception as e:
        logger.critical(f"An unhandled error occurred in main: {e}\n{traceback.format_exc()}")
    finally:
        for handler in logger.handlers[:]:
            try:
                handler.flush()
                handler.close()
                logger.removeHandler(handler)
            except Exception as e:
                print(f"Error closing log handler: {e}")

        # Automatically open the log file for review
        try:
            if os.path.exists(log_filename):
                print(f"Attempting to open log file: {log_filename}")
                if os.name == 'nt':
                    os.startfile(log_filename)
                elif os.uname().sysname == 'Darwin':
                    subprocess.run(['open', log_filename])
                else:
                    subprocess.run(['xdg-open', log_filename])
            else:
                print(f"Log file not found: {log_filename}")
        except Exception as e:
            print(f"Error opening log file {log_filename}: {e}")

        # Automatically open the Excel file for review
        try:
            if os.path.exists(excel_filename):
                print(f"Attempting to open Excel file: {excel_filename}")
                if os.name == 'nt':
                    os.startfile(excel_filename)
                elif os.uname().sysname == 'Darwin':
                    subprocess.run(['open', excel_filename])
                else:
                    subprocess.run(['xdg-open', excel_filename])
            else:
                print(f"Excel file not found: {excel_filename}")
        except Exception as e:
            print(f"Error opening Excel file {excel_filename}: {e}")