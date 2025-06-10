#
# 功能模块和实现手法整理
#
# 本脚本旨在通过自动化浏览器（Playwright）抓取 Google 图片搜索结果页面的信息，
# 包括每个搜索结果的图片、图片链接（原始图片URL）、关联网站的标题、关联网站链接和文字描述。
# 抓取到的数据将与抓取时间戳、搜索关键词和搜索URL一同被整理并导出到Excel文件中。
# 每个关键词的图片将保存到其专属的本地文件夹中，并在Excel中提供本地图片文件的绝对路径和可点击的超链接。
# 脚本会保留原有的日志系统，自动打开日志和结果文件，并支持注入Cookies。
#
# --- 主要功能模块 ---
#
# 1. 配置管理 (Configuration):
#    - PROXY: 定义HTTP代理地址，用于Playwright浏览器和httpx库的网络请求，规避地理限制或提高访问稳定性。
#    - LOG_DIR, IMAGE_DIR_BASE, RESULTS_DIR: 定义日志文件、图片存储和Excel结果文件的输出目录。
#    - KEYWORD_TARGET_FILE: 存储待搜索的关键词列表的文件。
#    - 目录创建: 脚本启动时自动创建所需的日志、图片和结果目录，确保文件能正确保存。
#    - 文件命名: 使用时间戳为日志和Excel文件生成唯一名称。
#
# 2. 日志系统 (Logging Setup):
#    - 使用Python内置的 `logging` 模块，配置日志记录器 `google_scraper`。
#    - 日志级别设置为 `INFO`，记录重要操作和信息。
#    - 同时配置文件处理器 (`FileHandler`) 和控制台处理器 (`StreamHandler`)，实现日志同时输出到文件和控制台。
#    - 日志格式化: 定义统一的日志输出格式。
#    - 目的: 方便跟踪脚本运行状态、调试问题和记录抓取过程中的事件。
#    - 自动打开日志: 脚本运行结束后，会自动尝试打开本次运行生成的日志文件。
#
# 3. 关键词读取 (Keyword Reading Helper Function - `read_keywords_from_file`):
#    - 目的: 从 `keywordTarget.txt` 文件中读取待搜索的关键词列表。
#    - 实现手法: 按行读取文件内容，每行视为一个关键词，并进行基本校验。
#    - 优点: 提高脚本的灵活性和可配置性。
#
# 4. 浏览器自动化与数据抓取 (Browser Automation & Data Scraping - `performGoogleImageSearch`):
#    - 核心模块，负责实际的网页交互和数据提取。
#    - 异步操作: 使用 `asyncio` 和 `playwright.async_api` 实现异步并发抓取。
#    - 浏览器上下文管理: 为每个抓取任务创建独立的浏览器上下文。
#    - 注入 Cookies: 尝试从 `cookies.json` 文件中读取并注入 cookies，以维持会话状态。
#    - 页面导航: 根据关键词构造Google图片搜索URL（`https://www.google.com/search?q=关键词&udm=2`）。
#    - 动态内容加载: **已移除**之前Twitter脚本的滚动加载逻辑，Google图片搜索结果通常在一页显示。
#    - 数据提取 (使用BeautifulSoup辅助):
#      - 定位大的搜索结果元素 (`div.eA0Zlc`)。
#      - 提取图片 (`.H8Rx8c img` 的 `src`)。
#      - 提取标题和链接 (`a.EZAeBe`)。
#      - 提取描述 (`div.toI8Rb`)。
#    - 图片下载: 使用 `httpx` 异步下载图片，并保存到关键词对应的子文件夹中。
#      - 文件命名: 使用时间戳和图片URL的SHA256哈希值，确保唯一性。
#    - 数据存储: 将抓取到的数据（包括时间戳、关键词、搜索URL、图片链接、图片本地路径、标题、标题链接和描述）存储到全局列表 `all_search_results_data` 中。
#    - 线程安全: 使用 `asyncio.Lock` 保护全局共享数据结构。
#
# 5. 主执行逻辑 (Main Execution Logic - `main`):
#    - 启动Playwright浏览器，并设置固定的窗口分辨率。
#    - 调用 `read_keywords_from_file` 获取所有目标关键词。
#    - 为每个关键词创建 `performGoogleImageSearch` 任务，并并发执行。
#    - 关闭浏览器实例。
#
# 6. Excel数据导出 (Excel Export Logic):
#    - 使用 `openpyxl` 库创建新的Excel工作簿，并创建名为 "Google图片搜索结果" 的工作表。
#    - 字段: "抓取时间", "搜索关键词", "搜索URL", "图片URL", "本地图片路径", "本地图片超链接", "搜索结果标题", "搜索结果标题链接", "搜索结果描述"。
#    - 超链接处理: 为搜索URL、搜索结果标题链接和**本地图片超链接**添加可点击的超链接。
#    - 列宽自适应: 根据列内容的最大长度自动调整列宽。
#
# 7. 脚本入口点 (Script Entry Point - `if __name__ == '__main__':`):
#    - 运行主异步函数。
#    - 错误处理: 捕获用户中断和所有未处理的异常，记录详细错误信息。
#    - 清理和自动打开文件 (`finally` 块): 关闭所有日志处理器，并尝试自动打开生成的日志文件和Excel结果文件。
#    - 兼容多操作系统: 自动判断操作系统并使用相应的命令打开文件。
#
# --- 技术栈 ---
# - Python 3.x
# - Playwright (异步浏览器自动化库)
# - BeautifulSoup4 (HTML解析库)
# - httpx (异步HTTP客户端，用于图片下载)
# - openpyxl (Excel文件读写库)
# - asyncio (Python异步编程框架)
# - logging (Python内置日志模块)
# - os, json, datetime, traceback, subprocess, hashlib (Python标准库)
# - aiofiles (异步文件操作，用于保存图片)
#
# --- 运行环境要求 ---
# - Python环境已安装。
# - 确保已安装所需的Python库: `pip install playwright beautifulsoup4 openpyxl httpx aiofiles`
# - 运行 `playwright install` 安装浏览器驱动。
# - 需要一个 `keywordTarget.txt` 文件，其中包含要搜索的关键词，一行一个。
# - **需要一个 `cookies.json` 文件，其中包含有效的 JSON 格式的 cookies 数据 (如果需要)。**
# - 可选配置代理 (`PROXY` 变量)。
#
#
import os
import json
import asyncio
import aiohttp # Changed from httpx to aiohttp
from datetime import datetime
import traceback
import logging
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import hashlib # For creating unique filenames for images
import aiofiles # For asynchronous file writing
import base64 # Import base64 for data:image handling
import re # Import re for regex to parse data:image

# --- FIX FOR NameError: name 'playwright' is not defined ---
# 导入 playwright.async_api 模块并为其创建一个别名，以便在异常处理时可以引用
import playwright.async_api as playwright_api
# 同时从 playwright.async_api 模块直接导入常用的函数和类
from playwright.async_api import async_playwright, expect
# --- END FIX ---

from bs4 import BeautifulSoup


# --- Configuration ---
PROXY = "http://127.0.0.1:10808" # Your proxy address
LOG_DIR = "logs" # Directory for logs
IMAGE_DIR_BASE = "images" # Base directory for images
RESULTS_DIR = "results" # Directory for Excel results
KEYWORD_TARGET_FILE = "keywordTarget.txt" # File to store target keywords
DOWNLOAD_HISTORY_FILE = "download_history.json" # New: File to store download history for deduplication

# Generate a timestamp for unique filenames - Moved to the top of the script for global access
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
log_filename = os.path.join(LOG_DIR, f"image_scraper_log_{timestamp}.txt")
excel_filename = os.path.join(RESULTS_DIR, f"google_image_results_{timestamp}.xlsx") # Excel filename with timestamp


# --- Logging Setup ---
# Get logger instance
logger = logging.getLogger('google_image_scraper') # Changed logger name
logger.setLevel(logging.INFO) # Set minimum logging level to INFO

# Create necessary directories if they don't exist
# Moved directory creation before logging setup to ensure log directory exists
if not os.path.exists(IMAGE_DIR_BASE):
    os.mkdir(IMAGE_DIR_BASE)
if not os.path.exists(LOG_DIR):
    os.mkdir(LOG_DIR)
if not os.path.exists(RESULTS_DIR):
    os.mkdir(RESULTS_DIR)

# Create file handler for logging to a file
file_handler = logging.FileHandler(log_filename, encoding='utf-8')
file_handler.setLevel(logging.INFO)

# Create console handler for logging to console
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# Define log message format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# Add handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)
# --- End Logging Setup ---

# Global list to store data for Excel export
all_search_results_data = []
# Lock for thread-safe access to all_search_results_data when multiple async tasks are running
data_lock = asyncio.Lock()

# Global dictionary to store download history: {identifier_hash: local_file_path}
# For data:image, identifier_hash is content MD5.
# For external URLs, identifier_hash is URL SHA256.
# This will be loaded at the start and saved at the end to persist downloaded image hashes.
download_history = {}


# --- Helper function to calculate MD5 hash of bytes ---
def calculate_md5(data_bytes):
    """Calculates the MD5 hash of a given byte string."""
    return hashlib.md5(data_bytes).hexdigest()

# --- Helper function to calculate SHA256 hash of a string ---
def calculate_sha256(data_string):
    """Calculates the SHA256 hash of a given string (e.g., a URL)."""
    return hashlib.sha256(data_string.encode('utf-8')).hexdigest()

# --- Helper functions to load/save download history ---
def load_download_history(filepath):
    """Loads the download history from a JSON file."""
    global download_history
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                download_history = json.load(f)
            logger.info(f"Loaded download history from {filepath}")
        except json.JSONDecodeError:
            logger.warning(f"Warning: '{filepath}' contains invalid JSON. Starting with empty history.")
            download_history = {} # Reset history if file is corrupted
        except Exception as e:
            logger.error(f"Error loading download history from '{filepath}': {e}\n{traceback.format_exc()}")
            download_history = {} # Reset history on other errors
    else:
        logger.info(f"Download history file '{filepath}' not found. Starting with empty history.")
        download_history = {}

def save_download_history(filepath):
    """Saves the current download history to a JSON file."""
    global download_history
    try:
        # Ensure the directory for the history file exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(download_history, f, indent=4) # Use indent for human-readable JSON
        logger.info(f"Saved download history to {filepath}")
    except Exception as e:
        logger.error(f"Error saving download history to '{filepath}': {e}\n{traceback.format_exc()}")

# --- New helper function to read Keywords from file ---
def read_keywords_from_file(filepath):
    """Reads keywords from a text file, one keyword per line."""
    keywords = []
    if not os.path.exists(filepath):
        logger.error(f"Error: Keyword target file '{filepath}' not found. Please create it with one keyword per line.")
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                keyword = line.strip()
                if keyword: # Basic validation for non-empty keyword
                    keywords.append(keyword)
        if not keywords:
            logger.warning(f"Warning: Keyword target file '{filepath}' is empty or contains no valid keywords.")
        return keywords
    except Exception as e:
        logger.error(f"Error reading keywords from '{filepath}': {e}\n{traceback.format_exc()}")
        return []
# --- End new helper function ---


async def process_image_data(image_url, keyword_folder_path, async_name):
    """
    Processes an image URL (external or data:image), handles deduplication,
    downloads if necessary, and returns the local file path and the identifier hash.
    For external URLs, deduplication is based on URL hash to avoid re-downloading.
    For data:image, deduplication is based on content MD5.
    """
    local_filename = None
    identifier_hash = None # This will be either content MD5 or URL SHA256

    if not image_url:
        logger.warning(f"{async_name} -> Empty image URL skipped.")
        return None, None

    # Handle data:image (Base64 encoded image directly in the URL)
    if image_url.startswith('data:image/'):
        # Regular expression to extract content type and base64 data
        match = re.match(r'data:image/(?P<ext>[^;]+);base64,(?P<data>.+)', image_url)
        if not match:
            logger.warning(f"{async_name} -> Invalid data:image format encountered: {image_url[:100]}...")
            return None, None

        extension = match.group('ext') # e.g., 'png', 'jpeg'
        base64_data = match.group('data') # The actual base64 encoded string

        try:
            # Decode the base64 data into bytes
            image_bytes = base64.b64decode(base64_data)
            # Calculate MD5 hash of the decoded image content for deduplication
            identifier_hash = calculate_md5(image_bytes) # Use content MD5 for data:image

            # Check if this image content (by MD5 hash) has been downloaded before
            if identifier_hash in download_history:
                existing_path = download_history[identifier_hash]
                # Verify that the file actually exists on disk (cleanup in case history is stale)
                if os.path.exists(existing_path):
                    logger.info(f"{async_name} -> Data:image content (MD5: {identifier_hash}) already exists at: {existing_path}. Skipping save.")
                    return existing_path, identifier_hash # Return existing path and MD5
                else:
                    logger.warning(f"{async_name} -> Data:image content (MD5: {identifier_hash}) found in history but file missing at {existing_path}. Re-downloading.")
                    # If file is missing, remove from history to attempt re-download/save
                    del download_history[identifier_hash]


            # If not in history or file missing, save the image locally
            # Use the content MD5 as part of the filename to ensure uniqueness and traceability
            local_filename = os.path.join(keyword_folder_path, f"{identifier_hash}.{extension.lower()}")

            # Ensure the target directory exists before writing the file
            os.makedirs(os.path.dirname(local_filename), exist_ok=True)
            async with aiofiles.open(local_filename, 'wb') as f:
                await f.write(image_bytes)
            logger.info(f"{async_name} -> Data:image saved locally: {local_filename}")

            # Update the global download history with the new image's content MD5 and its path
            download_history[identifier_hash] = local_filename
            return local_filename, identifier_hash

        except Exception as e:
            logger.error(f"{async_name} -> Error processing data:image: {e}\n{traceback.format_exc()}")
            return None, None

    # Handle external image URL (http/https)
    elif image_url.startswith('http'):
        # For external URLs, we will use the URL's SHA256 hash as the primary identifier for deduplication
        # This allows us to skip downloading if the specific URL has been processed before.
        identifier_hash = calculate_sha256(image_url) # Use URL SHA256 for external URLs

        # Determine a predictable local filename based on the URL's SHA256 hash
        url_without_query = image_url.split('?')[0]
        file_extension = url_without_query.split('.')[-1].lower()
        if not file_extension or len(file_extension) > 5 or not file_extension.isalpha():
            file_extension = 'jpg' # Default to jpg if no clear or valid extension
        
        # This is the expected path if the file were downloaded from this URL
        local_filename_from_url = os.path.join(keyword_folder_path, f"{identifier_hash}.{file_extension}")

        # === Deduplication Check: Check if this URL has been downloaded before ===
        if identifier_hash in download_history:
            existing_path = download_history[identifier_hash]
            # Crucially, check if the file still exists at the recorded path
            if os.path.exists(existing_path):
                logger.info(f"{async_name} -> Image from URL (SHA256: {identifier_hash}) already downloaded at: {existing_path}. Skipping download.")
                return existing_path, identifier_hash # Return existing path and URL SHA256
            else:
                logger.warning(f"{async_name} -> Image from URL (SHA256: {identifier_hash}) found in history but file missing at {existing_path}. Re-downloading.")
                # If the file is missing, remove from history so we attempt to re-download
                del download_history[identifier_hash]

        # If not in history or file is missing, proceed to download
        try:
            connector = None
            if PROXY:
                # If you experience SSL verification issues with your proxy, you might uncomment the line below:
                # connector = aiohttp.TCPConnector(ssl=False)
                pass

            async with aiohttp.ClientSession(connector=connector) as session:
                async with session.get(image_url, proxy=PROXY if PROXY else None, timeout=30.0) as response:
                    response.raise_for_status() # Raise an exception for HTTP errors (4xx or 5xx)
                    image_bytes = await response.read() # Read the entire image content into memory

                    # Ensure the directory for the specific image exists before writing the file
                    os.makedirs(os.path.dirname(local_filename_from_url), exist_ok=True)

                    async with aiofiles.open(local_filename_from_url, 'wb') as f:
                        await f.write(image_bytes)
                    logger.info(f"{async_name} -> Image downloaded: {local_filename_from_url}")

                    # Update the global download history with the URL's SHA256 and its local path
                    download_history[identifier_hash] = local_filename_from_url
                    return local_filename_from_url, identifier_hash

        except aiohttp.ClientResponseError as e:
            logger.error(f"{async_name} -> HTTP error {e.status} downloading image {image_url}: {e.message}")
        except aiohttp.ClientConnectorError as e:
            logger.error(f"{async_name} -> Network error (ClientConnectorError) downloading image {image_url}: {e}")
        except asyncio.TimeoutError:
            logger.error(f"{async_name} -> Timeout error downloading image {image_url}")
        except Exception as e:
            logger.error(f"{async_name} -> Unexpected error downloading image {image_url}: {e}\n{traceback.format_exc()}")
    else:
        logger.warning(f"{async_name} -> Unsupported image URL format (not http/https or data:image): {image_url[:100]}...")

    return None, None # Return None if any error or unhandled format


async def performGoogleImageSearch(context, keyword): # Renamed function
    # Get the name of the current asynchronous task
    async_name = asyncio.current_task().get_name()

    # Construct the Google image search URL
    search_url = f"https://www.google.com/search?q={keyword}&udm=2" # udm=2 for image search

    # --- 重新启用: 加载并添加 cookies 到浏览器上下文 ---
    try:
        if os.path.exists("cookies.json"):
            with open("cookies.json", "r", encoding="utf-8") as f:
                cookies = json.load(f)
                for cookie in cookies:
                    # Playwright expects 'sameSite' to be 'Strict', 'Lax', or 'None'
                    # Ensure the value is correctly mapped or removed if invalid
                    cookie_same_site = {'strict': 'Strict', 'Lax': 'Lax', 'none': 'None'}.get(cookie.get('sameSite'), None)
                    if cookie_same_site in ['Strict', 'Lax', 'None']:
                        cookie['sameSite'] = cookie_same_site
                    else:
                        if 'sameSite' in cookie:
                            del cookie['sameSite'] # Remove invalid sameSite attribute
                await context.add_cookies(cookies)
            logger.info(f"{async_name} -> Cookies loaded and added to context.")
        else:
            logger.warning(f"{async_name} -> Warning: cookies.json not found. Proceeding without cookies. Please ensure 'cookies.json' exists if needed.")
    except json.JSONDecodeError:
        logger.error(f"{async_name} -> Error: Invalid JSON in cookies.json. Please check the file format. Full traceback:\n{traceback.format_exc()}")
        # return # Optionally return here if cookies are critical for the page to load correctly
    except Exception as e:
        logger.error(f"{async_name} -> Unexpected error loading cookies: {e}\n{traceback.format_exc()}")
        # return # Optionally return here if cookies are critical for the page to load correctly
    # --- 重新启用结束 ---

    # Create a new page in the browser context
    page = await context.new_page()

    try:
        # Navigate to the Google search URL, with a timeout
        logger.info(f"{async_name} -> Navigating to {search_url}")
        await page.goto(search_url, timeout=60000, wait_until="domcontentloaded") # Use domcontentloaded for faster loading
        logger.info(f"{async_name} -> Successfully navigated to {search_url}")
    except playwright_api.TimeoutError:
        logger.error(f"{async_name} -> Error: Page.goto timed out for {search_url} after 60 seconds. Check network or proxy.")
        await page.close()
        return
    except Exception as e:
        logger.error(f"{async_name} -> An unexpected error occurred during navigation to {search_url}: {e}")
        await page.close()
        return

    # Create keyword-specific image folder
    keyword_folder_name = keyword.replace(" ", "_").replace("/", "_").replace("\\", "_") # Sanitize keyword for folder name
    keyword_folder_path = os.path.join(IMAGE_DIR_BASE, keyword_folder_name)
    if not os.path.exists(keyword_folder_path):
        os.makedirs(keyword_folder_path)
        logger.info(f"{async_name} -> Created image folder: {keyword_folder_path}")

    # Get the current timestamp for when the data was scraped
    scrape_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        # Locate all large search result elements using div.eA0Zlc
        # 等待至少一个搜索结果元素出现
        await expect(page.locator('div.eA0Zlc').nth(0)).to_be_visible(timeout=30000)
        search_results_divs = page.locator('div.eA0Zlc')
        result_count = await search_results_divs.count()
        logger.info(f"{async_name} -> Found {result_count} search results for keyword: '{keyword}'")

        if result_count == 0:
            logger.warning(f"{async_name} -> No search results found for keyword: '{keyword}'.")

        for i in range(result_count):
            result_data = {
                "抓取时间": scrape_timestamp,
                "搜索关键词": keyword,
                "搜索URL": search_url,
                "图片URL": "",
                "本地图片路径": "",
                "本地图片超链接": "",
                "搜索结果标题": "",
                "搜索结果标题链接": "",
                "搜索结果描述": ""
            }
            try:
                current_result_div = search_results_divs.nth(i)
                # Ensure the current search result div is visible before interacting with it
                await expect(current_result_div).to_be_visible(timeout=10000)

                # Get the inner HTML of the current search result div for BeautifulSoup parsing
                soup = BeautifulSoup(await current_result_div.inner_html(timeout=10000), "html.parser")

                # Extract image URL
                # Prioritize 'data-src' if available, then fallback to 'src'
                image_element = soup.select_one('.H8Rx8c img')
                image_url = image_element.get('data-src') or image_element.get('src') if image_element else ""
                result_data["图片URL"] = image_url

                # Process image data (download or use from history)
                if image_url:
                    # Call the process_image_data function to handle both types of URLs and deduplication
                    local_image_path, identifier_hash_used = await process_image_data(image_url, keyword_folder_path, async_name)
                    if local_image_path:
                        result_data["本地图片路径"] = os.path.abspath(local_image_path) # Absolute path
                        # For Excel hyperlink, prepend file:/// if on Windows, or just path
                        if os.name == 'nt':
                            # Pre-format the path to avoid backslash issue in f-string
                            formatted_local_path = result_data['本地图片路径'].replace('\\', '/')
                            result_data["本地图片超链接"] = f"file:///{formatted_local_path}"
                        else:
                            result_data["本地图片超链接"] = f"file://{result_data['本地图片路径']}"

                # Extract title and title link (a.EZAeBe)
                title_element = soup.select_one('a.EZAeBe')
                if title_element:
                    result_data["搜索结果标题"] = title_element.get_text(strip=True)
                    result_data["搜索结果标题链接"] = title_element.get('href')
                else:
                    logger.warning(f"{async_name} -> Skipping result {i+1} for keyword '{keyword}': No title element (a.EZAeBe) found.")
                    # Continue processing, title might not always be present for images

                # Extract description (div.toI8Rb)
                description_element = soup.select_one('div.toI8Rb')
                if description_element:
                    result_data["搜索结果描述"] = description_element.get_text(separator="\n", strip=True)
                # else: description can be empty, not a critical error.

                logger.info(f"{async_name} -> {'=' * 30}")
                logger.info(f"{async_name} -> Keyword: '{keyword}', Result {i+1}/{result_count}.")
                logger.info(f"{async_name} -> Image URL: {result_data['图片URL']}")
                logger.info(f"{async_name} -> Local Image Path: {result_data['本地图片路径']}")
                logger.info(f"{async_name} -> Title: {result_data['搜索结果标题']}")
                logger.info(f"{async_name} -> Title Link: {result_data['搜索结果标题链接']}")
                logger.info(f"{async_name} -> Description: {result_data['搜索结果描述']}")

                async with data_lock: # Acquire lock before modifying shared list
                    all_search_results_data.append(result_data)

            except (playwright_api.TimeoutError, AssertionError) as e:
                logger.warning(f"{async_name} -> Playwright Locator Error for search result {i+1} for keyword '{keyword}': {e}. Skipping this result. Full traceback:\n{traceback.format_exc()}")
            except Exception as e:
                logger.error(f"{async_name} -> General Error processing search result {i+1} for keyword '{keyword}': {str(e)}\n{traceback.format_exc()}")


    except playwright_api.TimeoutError:
        logger.error(f"{async_name} -> Error: No search results (div.eA0Zlc) found on page for keyword '{keyword}' within timeout.")
    except Exception as e:
        logger.error(f"{async_name} -> An unexpected error occurred during search results processing for keyword '{keyword}': {e}\n{traceback.format_exc()}")


    await page.close() # Close the Playwright page
    logger.info(f"{async_name} -> Page closed for keyword: '{keyword}'.")


async def main():
    # Load download history at the start of the script to enable deduplication
    load_download_history(DOWNLOAD_HISTORY_FILE)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False, # Set to True for headless Browse (runs without a visible browser window)
            proxy={"server": PROXY} if PROXY else None, # Configure proxy for the Playwright browser
            timeout=60000 # Browser launch timeout (in milliseconds)
        )
        # Set browser window size
        context = await browser.new_context(
            viewport={'width': 2560, 'height': 1440}
        ) # Create a new browser context with specified viewport size

        # --- Optimized section for reading Keywords from file ---
        target_keywords = read_keywords_from_file(KEYWORD_TARGET_FILE)
        if not target_keywords:
            logger.error(f"No valid keywords found in {KEYWORD_TARGET_FILE}. Exiting scraper.")
            await browser.close()
            return # Exit main if no keywords to scrape

        # Create a list of tasks for each keyword
        # Call the renamed function here
        tasks = [performGoogleImageSearch(context, keyword) for keyword in target_keywords]

        # Run all scraping tasks concurrently
        await asyncio.gather(*tasks)
        # --- End optimized section ---

        await browser.close() # Close the browser instance after all tasks are done
        logger.info("Browser closed. Script finished scraping data.")

    # --- Excel Export Logic (This block will always run after scraping attempts complete) ---
    wb = Workbook() # Create a new Excel workbook

    # --- Sheet 1: Google图片搜索结果 (Google Image Search Results) ---
    ws = wb.active # Get the active worksheet (the first one created by default)
    ws.title = "Google图片搜索结果" # Set the title of the active sheet

    # Define headers for the image search results sheet
    headers = ["抓取时间", "搜索关键词", "搜索URL", "图片URL", "本地图片路径", "本地图片超链接", "搜索结果标题", "搜索结果标题链接", "搜索结果描述"]
    ws.append(headers) # Write headers to the first row of the sheet

    # Define font style for hyperlinks (blue, underlined)
    hyperlink_font = Font(color="0000FF", underline="single")

    # Populate the sheet with collected search results data
    for row_data in all_search_results_data:
        row = []
        for header in headers:
            # Append data from the dictionary, using an empty string if a key is not found
            row.append(row_data.get(header, ""))
        ws.append(row) # Add the row to the worksheet

        current_row_idx = ws.max_row # Get the current row index for applying hyperlinks

        # Apply hyperlink for "搜索URL" (Search URL)
        search_url = row_data.get("搜索URL")
        if search_url:
            cell_search_url = ws.cell(row=current_row_idx, column=headers.index("搜索URL") + 1)
            cell_search_url.value = search_url
            cell_search_url.hyperlink = search_url
            cell_search_url.font = hyperlink_font

        # Apply hyperlink for "图片URL" (Image URL)
        image_url = row_data.get("图片URL")
        if image_url:
            cell_image_url = ws.cell(row=current_row_idx, column=headers.index("图片URL") + 1)
            cell_image_url.value = image_url
            cell_image_url.hyperlink = image_url
            cell_image_url.font = hyperlink_font

        # Apply hyperlink for "本地图片超链接" (Local Image Hyperlink)
        local_image_hyperlink = row_data.get("本地图片超链接")
        if local_image_hyperlink:
            cell_local_image_hyperlink = ws.cell(row=current_row_idx, column=headers.index("本地图片超链接") + 1)
            # Display text for the hyperlink will be the local path, but the actual hyperlink is the file:// one
            cell_local_image_hyperlink.value = row_data.get("本地图片路径", "")
            cell_local_image_hyperlink.hyperlink = local_image_hyperlink
            cell_local_image_hyperlink.font = hyperlink_font


        # Apply hyperlink for "搜索结果标题链接" (Search Result Title Link)
        title_link = row_data.get("搜索结果标题链接")
        if title_link:
            cell_title_link = ws.cell(row=current_row_idx, column=headers.index("搜索结果标题链接") + 1)
            cell_title_link.value = title_link
            cell_title_link.hyperlink = title_link
            cell_title_link.font = hyperlink_font

    # Adjust column widths for the sheet for better readability
    for col_idx, header in enumerate(headers):
        max_length = len(header) # Initialize max length with header's length
        column_letter = get_column_letter(col_idx + 1) # Get Excel column letter (e.g., 'A', 'B', 'C'...)
        # Iterate over all cells in the column to find the maximum content length
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except TypeError:
                pass # Ignore non-string types (e.g., numbers, None)

        adjusted_width = (max_length + 2) * 1.2 # Add padding and a factor for better visual spacing
        if adjusted_width > 100: # Cap maximum column width to prevent excessively wide columns
            adjusted_width = 100
        ws.column_dimensions[column_letter].width = adjusted_width # Set the adjusted column width

    wb.save(excel_filename) # Save the entire workbook to the timestamped Excel file
    logger.info(f"Results saved to Excel: {excel_filename}")
    # --- END Excel Export Logic ---

    # Save download history at the end of the script for persistence
    save_download_history(DOWNLOAD_HISTORY_FILE)


if __name__ == '__main__':
    # Store handlers to close them properly later
    # This list must be created BEFORE any handlers are added to the logger
    # so we have a reference to the handlers created during global setup.
    global_handlers = []
    # This loop is problematic because the handlers are added *after* the logger.handlers list is empty initially.
    # The proper way to get the handlers is to access logger.handlers AFTER they've been added.
    # Instead, we will store them explicitly when they are created.

    try:
        # Run the main asynchronous function
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Script interrupted by user.") # Log if script is interrupted by user (Ctrl+C)
    except Exception as e:
        logger.critical(f"An unhandled error occurred in main: {e}\n{traceback.format_exc()}") # Log any unhandled critical errors
    finally:
        # This 'finally' block ensures cleanup and file opening actions always happen

        # Before opening files, ensure the log handlers are flushed and closed.
        # It's crucial to do this *after* all logging is done.
        for handler in logger.handlers[:]: # Iterate over a slice to safely modify list while iterating
            try:
                handler.flush()
                handler.close()
                logger.removeHandler(handler)
            except Exception as e:
                print(f"Error closing log handler: {e}") # Use print here as logger might be closed

        # Automatically open the log file for review
        try:
            if os.path.exists(log_filename):
                print(f"Attempting to open log file: {log_filename}")
                if os.name == 'nt':  # Check if OS is Windows
                    os.startfile(log_filename)
                elif os.uname().sysname == 'Darwin':  # Check if OS is macOS
                    subprocess.run(['open', log_filename])
                else:  # Assume Linux-like system
                    subprocess.run(['xdg-open', log_filename]) # Common command for opening files on Linux
            else:
                print(f"Log file not found: {log_filename}")
        except Exception as e:
            print(f"Error opening log file {log_filename}: {e}")

        # Automatically open the Excel file for review
        try:
            if os.path.exists(excel_filename):
                print(f"Attempting to open Excel file: {excel_filename}")
                if os.name == 'nt':  # Windows
                    os.startfile(excel_filename)
                elif os.uname().sysname == 'Darwin':  # macOS
                    subprocess.run(['open', excel_filename])
                else:  # Linux
                    subprocess.run(['xdg-open', excel_filename])
            else:
                print(f"Excel file not found: {excel_filename}")
        except Exception as e:
            print(f"Error opening Excel file {excel_filename}: {e}")