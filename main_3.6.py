# -*- coding: utf-8 -*-
"""
Civitai图片爬虫主流程：只用核心滚动代码，抓取所有图片信息并保存到Excel
"""

import os
import json
import asyncio
import aiohttp
from datetime import datetime
import traceback
import logging
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import hashlib
import aiofiles
import re
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
import playwright._impl._errors
import time
import shutil # 导入 shutil 模块用于文件复制

# --- 配置 ---
PROXY = "http://127.0.0.1:10808"
TARGET_URL = "https://civitai.com/images?tags=4"
LOG_DIR = "logs"
RESULTS_DIR = "results_civitai"
IMAGE_DIR_BASE = "images_civitai"
KEYWORD_TARGET_FILE = "urlTarget.txt"
DOWNLOAD_HISTORY_FILE = "download_history_civitai.json" # Original MD5-based history
HISTORY_IMG_URL_FILE = "history_img_url_history.xlsx" # New URL/Path based history
HISTORY_MD5_DIR = "historyImgMD5" # 新增：存放带时间戳的历史记录副本的目录

timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
log_filename = os.path.join(LOG_DIR, f"civitai_scraper_log_{timestamp}.txt")
excel_filename = os.path.join(RESULTS_DIR, f"civitai_image_results_{timestamp}.xlsx")

# --- 日志配置 ---
logger = logging.getLogger('civitai_scraper')
logger.setLevel(logging.INFO)
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
if not os.path.exists(RESULTS_DIR):
    os.makedirs(RESULTS_DIR)
if not os.path.exists(IMAGE_DIR_BASE):
    os.makedirs(IMAGE_DIR_BASE)
if not os.path.exists(HISTORY_MD5_DIR): # 创建新的历史记录副本目录
    os.makedirs(HISTORY_MD5_DIR)
    logger.info(f"Created history copy directory: {HISTORY_MD5_DIR}")

file_handler = logging.FileHandler(log_filename, encoding='utf-8')
console_handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# --- 工具函数 ---
def calculate_md5(data_bytes):
    return hashlib.md5(data_bytes).hexdigest()

def calculate_url_md5(url):
    """Calculates MD5 for a given URL string, used for consistent naming."""
    return hashlib.md5(url.encode('utf-8')).hexdigest()

# Global history for MD5-based downloaded content
download_history = {}
# Global history for URL-based downloaded content (new)
url_download_history = {}

# Locks for shared resources
download_history_lock = asyncio.Lock()
url_download_history_lock = asyncio.Lock()
all_search_results_data_lock = asyncio.Lock()

def load_download_history(filepath):
    global download_history
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                download_history = json.load(f)
            logger.info(f"Loaded download history from {filepath}")
        except Exception as e:
            logger.warning(f"Failed to load download history (MD5): {e}")
            download_history = {}
    else:
        logger.info(f"Download history file '{filepath}' not found. Starting with empty history.")
        download_history = {}

def save_download_history(filepath):
    global download_history
    try:
        dir_name = os.path.dirname(filepath)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(download_history, f, indent=4)
        logger.info(f"Saved download history to {filepath}")
    except Exception as e:
        logger.error(f"Error saving download history (MD5) to '{filepath}': {e}\n{traceback.format_exc()}")

def load_url_history(filepath):
    global url_download_history
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            for row_idx in range(2, ws.max_row + 1):
                row_data = {header[col_idx]: ws.cell(row=row_idx, column=col_idx+1).value for col_idx in range(len(header))}
                thumb_url = row_data.get("Thumbnail URL")
                orig_page_url = row_data.get("Original Page URL")
                local_path = row_data.get("Local Image Path")
                if thumb_url and orig_page_url and local_path:
                    # Using a combined key for uniqueness
                    key = f"{thumb_url}|{orig_page_url}"
                    url_download_history[key] = {
                        "local_path": local_path, # 这里加载进来已经是绝对路径了
                        "image_md5": row_data.get("MD5 (Content)") # This might be useful for verification if needed
                    }
            logger.info(f"Loaded URL download history from {filepath}")
        except Exception as e:
            logger.warning(f"Failed to load URL download history from {filepath}: {e}")
            url_download_history = {}
    else:
        logger.info(f"URL download history file '{filepath}' not found. Starting with empty history.")
        url_download_history = {}

def save_url_history(filepath):
    global url_download_history
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Download History"
        headers = ["Thumbnail URL", "Original Page URL", "Local Image Path", "MD5 (Content)"]
        ws.append(headers)

        for key, data in url_download_history.items():
            thumb_url, orig_page_url = key.split('|', 1) # Split only on the first '|'
            # Ensure local_path is saved with appropriate slashes for the OS,
            # though the hyperlink generation below will handle conversion for file:// URLs.
            local_path_for_excel = data["local_path"] # 这里已经是绝对路径了
            ws.append([thumb_url, orig_page_url, local_path_for_excel, data.get("image_md5", "")])

        for col_idx, header in enumerate(headers):
            max_length = len(header)
            column_letter = get_column_letter(col_idx + 1)
            for r_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=r_idx, column=col_idx + 1).value
                if cell_value:
                    cell_len = len(str(cell_value))
                    if cell_len > max_length:
                        max_length = cell_len
            adjusted_width = (max_length + 2) * 1.2
            if adjusted_width > 100:
                adjusted_width = 100
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        logger.info(f"Saved URL download history to {filepath}")
    except Exception as e:
        logger.error(f"Error saving URL download history to '{filepath}': {e}\n{traceback.format_exc()}")

def read_urls_from_file(filepath):
    urls = []
    if not os.path.exists(filepath):
        logger.error(f"Error: Target URL file '{filepath}' not found. 请创建并添加URL。")
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                url = line.strip()
                if url and url.startswith("http"):
                    urls.append(url)
        if not urls:
            logger.warning(f"Warning: Target URL file '{filepath}' is empty or contains no valid URLs.")
        return urls
    except Exception as e:
        logger.error(f"Error reading URLs from '{filepath}': {e}\n{traceback.format_exc()}")
        return []

# Helper to parse counts with 'k'
def parse_count_with_k(count_str):
    if not count_str:
        return 0
    count_str = count_str.strip().lower()
    if 'k' in count_str:
        try:
            return int(float(count_str.replace('k', '')) * 1000)
        except ValueError:
            return 0
    else:
        try:
            # Use regex to find only digits (and possibly a dot for k)
            match = re.search(r'[\d.]+', count_str)
            if match:
                return int(float(match.group(0)))
            return 0
        except ValueError:
            return 0

# --- 核心爬虫流程 ---

async def _navigate_and_setup_page(context, target_url, async_name):
    """
    负责页面导航和初始设置（如注入cookie）。
    Returns: playwright.Page or None if navigation fails.
    """
    try:
        if os.path.exists("cookies.json"):
            with open("cookies.json", "r", encoding="utf-8") as f:
                cookies = json.load(f)
                for cookie in cookies:
                    cookie_same_site = {'strict': 'Strict', 'Lax': 'Lax', 'none': 'None'}.get(cookie.get('sameSite'), None)
                    if cookie_same_site in ['Strict', 'Lax', 'None']:
                        cookie['sameSite'] = cookie_same_site
                    else:
                        if 'sameSite' in cookie:
                            del cookie['sameSite']
                await context.add_cookies(cookies)
            logger.info(f"{async_name} -> Cookies loaded and added to context.")
        else:
            logger.warning(f"{async_name} -> Warning: cookies.json not found. Proceeding without cookies.")
    except Exception as e:
        logger.error(f"{async_name} -> Error loading cookies: {e}")

    page = await context.new_page()
    try:
        logger.info(f"{async_name} -> Navigating to {target_url}")
        # Explicitly removed 'wait_until="networkidle"'
        await page.goto(target_url, timeout=60000, wait_until="domcontentloaded")
        # Add a short wait here to allow initial page content to render
        await asyncio.sleep(2) # 等待2秒，让页面初步加载完成
        logger.info(f"{async_name} -> Successfully navigated to {target_url}")
        return page
    except Exception as e:
        logger.error(f"{async_name} -> Error navigating to {target_url}: {e}")
        await page.close()
        return None

async def _extract_keyword(page, async_name):
    """
    负责从页面顶部输入框提取当前关键词。
    Returns: str
    """
    # Refined selector to target the search input specifically
    # Using first to resolve strict mode violation if multiple elements match
    keyword_input_element = page.get_by_placeholder("Search Civitai").first
    current_keyword = "N/A"
    try:
        # Check if the element is visible before trying to get its value
        if await keyword_input_element.is_visible():
            input_value = await keyword_input_element.get_attribute('value')
            if input_value:
                current_keyword = input_value
                logger.info(f"{async_name} -> Found keyword in input field: '{current_keyword}'")
    except playwright._impl._errors.Error as e:
        logger.warning(f"{async_name} -> Could not find or extract keyword due to Playwright error: {e}")
        # If it's a strict mode violation, try another more specific selector if possible
        # Or log the specific locators that caused the issue.
    except Exception as e:
        logger.warning(f"{async_name} -> Could not find or extract keyword: {e}")
    return current_keyword

async def _scroll_page(page):
    """
    负责执行页面滚动操作。
    """
    await page.evaluate("""
        document.querySelectorAll('*').forEach(function(el) {
            if (el.scrollHeight > el.clientHeight) el.scrollTop += 40;
        });
    """)
    #await asyncio.sleep(1) # 每次滚动后等待0.01秒，确保内容加载


def _extract_button_counts(button): # Changed from async def to def
    """
    辅助函数：从单个按钮元素中提取点赞、爱心、笑哭、伤心和打赏的数量。
    """
    like_count = 0
    heart_count = 0
    laugh_count = 0
    sad_count = 0
    tip_count = 0

    # Handle standard emoji buttons (like, heart, laugh, sad)
    label_span = button.find("span", class_="mantine-Button-label")
    if label_span:
        emoji_div = label_span.find("div", class_="mantine-Text-root")
        
        # Get all text from the span and then try to extract the number
        # This covers cases where the number is directly in the span, not in a separate div
        full_label_text = label_span.get_text(separator=' ', strip=True)
        
        # Use regex to find digits (and possibly 'k' for thousands)
        # [\d.]+ ensures we capture numbers like 1.2k, \d[\d\.]* ensures at least one digit and subsequent digits/dots
        match = re.search(r'(\d[\d\.]*[kK]?)', full_label_text)
        
        count = 0
        if match:
            count_str = match.group(1)
            count = parse_count_with_k(count_str)

        if emoji_div: # Ensure emoji div exists
            if "👍" in emoji_div.text:
                like_count = count
            elif "❤️" in emoji_div.text:
                heart_count = count
            elif "😂" in emoji_div.text:
                laugh_count = count
            elif "😢" in emoji_div.text:
                sad_count = count
    
    # Special handling for the tip button (it's a badge, not a simple button label)
    # This part remains unchanged as per your instruction
    tip_badge = button.find("div", class_=lambda x: x and "mantine-Badge-root" in x)
    if tip_badge:
        # Check for the lightning bolt SVG
        if tip_badge.find("svg", class_=lambda x: x and "tabler-icon-bolt" in x):
            tip_text_div = tip_badge.find("div", class_="mantine-Text-root") # The count is in this div
            if tip_text_div:
                tip_str = tip_text_div.text.strip()
                tip_count = parse_count_with_k(tip_str)

    return like_count, heart_count, laugh_count, sad_count, tip_count


async def _parse_card_container(card_container, image_download_queue, base_image_folder_path, target_url, current_keyword, processed_image_detail_urls):
    """
    负责解析单个图片卡片容器，提取所有相关信息并**将下载任务放入队列**。
    Returns: bool (True if a new task was added, False otherwise)
    """
    thumbnail_url = ""
    original_page_url = ""
    
    # Extract thumbnail URL and original page URL
    img_element = card_container.find("img", class_="EdgeImage_image__iH4_q")
    if img_element:
        thumbnail_url = img_element.get("src")
        parent_a = img_element.find_parent("a")
        if parent_a:
            original_page_url = parent_a.get("href")
            if original_page_url and not original_page_url.startswith("http"):
                original_page_url = f"https://civitai.com{original_page_url}"

    if not thumbnail_url or not thumbnail_url.startswith("http"):
        return False

    # De-duplication check for scraping process (not for download history)
    unique_key_for_scrape_tracking = thumbnail_url + "|" + original_page_url
    if unique_key_for_scrape_tracking in processed_image_detail_urls:
        return False # Already processed this card

    processed_image_detail_urls.add(unique_key_for_scrape_tracking)

    # --- Extract Button Counts ---
    like_count = 0
    heart_count = 0
    laugh_count = 0
    sad_count = 0
    tip_count = 0

    buttons_container = card_container.find("div", class_=lambda x: x and "flex items-center justify-center" in x and "p-2" in x and "gap-1" in x)
    
    if buttons_container:
        buttons = buttons_container.find_all("button", class_=lambda x: x and ("mantine-UnstyledButton-root" in x or "mantine-Button-root" in x))
        for button in buttons:
            l, h, la, s, t = _extract_button_counts(button)
            like_count += l
            heart_count += h
            laugh_count += la
            sad_count += s
            tip_count += t # Accumulate counts from all relevant buttons
    
    # Put download task into queue
    # The result data is initially created here with placeholder for local_image_path/md5
    # The actual download and MD5 calculation will happen in the consumer.
    result_data_template = {
        "抓取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "搜索URL": target_url,
        "缩略图URL": thumbnail_url,
        "本地缩略图路径": "", # To be filled by download worker
        "本地缩略图超链接": "", # To be filled by download worker
        "原始图片详情页链接": original_page_url,
        "点赞数": like_count,
        "爱心数": heart_count,
        "笑哭数": laugh_count,
        "伤心数": sad_count,
        "打赏数": tip_count,
        "关键词": current_keyword
    }
    
    await image_download_queue.put((thumbnail_url, original_page_url, base_image_folder_path, result_data_template))
    return True

async def image_downloader(image_download_queue, md5_analysis_queue):
    """
    Consumer task: Downloads images from the queue.
    """
    while True:
        task = await image_download_queue.get()
        if task is None: # Sentinel value to stop the worker
            image_download_queue.task_done()
            break

        image_url, original_page_url, base_folder_path, result_data_template = task
        local_filename = None
        image_bytes = None
        unique_url_key = f"{image_url}|{original_page_url}"

        # Determine file extension
        url_without_query = image_url.split('?')[0]
        file_extension = url_without_query.split('.')[-1].lower()
        if not file_extension or len(file_extension) > 5 or not file_extension.isalpha():
            file_extension = 'jpg'

        download_needed = True
        existing_path_from_history = None
        existing_md5_from_history = None

        # Check URL/Path history first to avoid re-downloading
        async with url_download_history_lock:
            if unique_url_key in url_download_history:
                existing_info = url_download_history[unique_url_key]
                existing_path_from_history = existing_info["local_path"]
                existing_md5_from_history = existing_info.get("image_md5")

                # Note: existing_path_from_history should already be absolute path due to previous fix
                if os.path.exists(existing_path_from_history): # Check against MD5 history path
                    logger.info(f"Image '{image_url}' found in URL history with local path '{existing_path_from_history}'. File exists. Skipping download.")
                    download_needed = False
                    local_filename = existing_path_from_history
                    image_md5 = existing_md5_from_history
                else:
                    logger.warning(f"Image '{image_url}' found in URL history, but local file '{existing_path_from_history}' is missing. Will re-download.")

        if download_needed and image_url.startswith('http'):
            try:
                async with aiohttp.ClientSession() as session:
                    async with session.get(image_url, proxy=PROXY if PROXY else None, timeout=30.0) as response:
                        response.raise_for_status()
                        image_bytes = await response.read()
                        logger.info(f"Image downloaded: {image_url}")
                        # Pass image_bytes to MD5 analyzer queue
                        await md5_analysis_queue.put((image_bytes, image_url, original_page_url, base_folder_path, file_extension, result_data_template))
            except Exception as e:
                logger.error(f"Error downloading image {image_url}: {e}")
        elif not download_needed:
            # If not downloaded (already in history and exists), directly add to results data
            if local_filename:
                # Fix for local path slashes for hyperlink
                abs_path = os.path.abspath(local_filename) # local_filename已经是绝对路径，这里再获取一次abspath不影响
                # Convert backslashes to forward slashes for file:// URLs
                local_image_hyperlink = f"file:///{abs_path.replace(os.sep, '/')}"

                result_data_template["本地缩略图路径"] = abs_path # Store as OS native path
                result_data_template["本地缩略图超链接"] = local_image_hyperlink
                # If we skipped download, use the MD5 from history
                result_data_template["image_md5"] = existing_md5_from_history # Add MD5 for consistency in results
                
                async with all_search_results_data_lock:
                    all_search_results_data.append(result_data_template)
        image_download_queue.task_done()

async def md5_analyzer(md5_analysis_queue):
    """
    Consumer task: Calculates MD5, saves image, and updates histories.
    """
    while True:
        task = await md5_analysis_queue.get()
        if task is None: # Sentinel value
            md5_analysis_queue.task_done()
            break

        image_bytes, image_url, original_page_url, base_folder_path, file_extension, result_data_template = task
        image_content_md5 = calculate_md5(image_bytes)
        
        # 确保 local_filename_base 是绝对路径，并且基于此计算出最终的绝对路径
        # base_folder_path 已经被 performCivitaiImageScrape 转换为绝对路径
        local_filename_base = os.path.join(base_folder_path, f"{image_content_md5}.{file_extension}")
        abs_local_filename = os.path.abspath(local_filename_base) # 再次确认是绝对路径

        unique_url_key = f"{image_url}|{original_page_url}"

        # Check MD5-based download history (secondary check for content de-duplication)
        # Note: download_history 现在也应该存储绝对路径
        existing_path_from_md5_history = None
        async with download_history_lock:
            if image_content_md5 in download_history:
                existing_path_from_md5_history = download_history[image_content_md5]
                if os.path.exists(existing_path_from_md5_history) and os.path.abspath(existing_path_from_md5_history) == abs_local_filename:
                    logger.info(f"Downloaded image content (MD5: {image_content_md5}) already exists at: {existing_path_from_md5_history}. Skipping save and using existing path.")
                else:
                    logger.warning(f"Image MD5 found in history, but file '{existing_path_from_md5_history}' is missing or path is outdated. Will re-save.")
            # If not in history, or if existing path was bad, the file will be saved to abs_local_filename

        # Save the image if it's new or was missing
        if not os.path.exists(abs_local_filename):
            os.makedirs(os.path.dirname(abs_local_filename), exist_ok=True)
            try:
                async with aiofiles.open(abs_local_filename, 'wb') as f:
                    await f.write(image_bytes)
                logger.info(f"Image saved: {abs_local_filename}")
            except Exception as e:
                logger.error(f"Error saving image {abs_local_filename}: {e}")
                md5_analysis_queue.task_done()
                continue
        else:
            logger.info(f"Image already exists at {abs_local_filename}. Skipping file write.")
        
        # Update both histories with the absolute path
        async with download_history_lock:
            download_history[image_content_md5] = abs_local_filename
        async with url_download_history_lock:
            url_download_history[unique_url_key] = {"local_path": abs_local_filename, "image_md5": image_content_md5}

        # Prepare result data for Excel
        local_image_hyperlink = f"file:///{abs_local_filename.replace(os.sep, '/')}"
        
        result_data_template["本地缩略图路径"] = abs_local_filename # Store as OS native path
        result_data_template["本地缩略图超链接"] = local_image_hyperlink
        result_data_template["image_md5"] = image_content_md5 # Add MD5 for consistency in results

        async with all_search_results_data_lock:
            all_search_results_data.append(result_data_template)
            
        md5_analysis_queue.task_done()

async def performCivitaiImageScrape(context, target_url, image_download_queue, max_scroll_attempts):
    # Declare local processed_image_detail_urls for this function's scope
    processed_image_detail_urls_local = set() 
    async_name = asyncio.current_task().get_name()
    
    page = await _navigate_and_setup_page(context, target_url, async_name)
    if not page:
        return

    civitai_image_folder_path = os.path.join(IMAGE_DIR_BASE, "downloaded_images")
    civitai_image_folder_path = os.path.abspath(civitai_image_folder_path) # 确保是绝对路径

    if not os.path.exists(civitai_image_folder_path):
        os.makedirs(civitai_image_folder_path)
        logger.info(f"{async_name} -> Created base image folder for Civitai: {civitai_image_folder_path}")

    current_keyword = await _extract_keyword(page, async_name)

    scroll_attempts = 0
    # Removed no_new_images_start_time and associated logic to stop scrolling based on new images

    while scroll_attempts < max_scroll_attempts:
        scroll_attempts += 1
        logger.info(f"{async_name} -> Scroll attempt {scroll_attempts}/{max_scroll_attempts}...")
        for _ in range(10): # 连续滚动五次
            await _scroll_page(page)
            # await asyncio.sleep(0.1) # Short delay between rapid scrolls if needed

        page_html = await page.content()
        soup = BeautifulSoup(page_html, "html.parser")
        target_div = soup.select_one("div.mx-auto.flex.justify-center.gap-4")
        if not target_div:
            logger.warning("目标div未找到，跳过本次循环")
            continue
        
        image_card_containers = target_div.find_all("div", class_=lambda x: x and "flex-col border" in x and "relative flex overflow-hidden" in x)

        newly_processed_this_scroll = 0 # Still track for logging/visibility if needed, but not for stopping
        for card_container in image_card_containers:
            # Pass the queue to _parse_card_container
            task_added = await _parse_card_container(card_container, image_download_queue, civitai_image_folder_path, target_url, current_keyword, processed_image_detail_urls_local)
            if task_added:
                newly_processed_this_scroll += 1
        
        logger.info(f"{async_name} -> Discovered {newly_processed_this_scroll} new image tasks this scroll.") # Optional: add this for more detailed logs

    await page.close()
    logger.info(f"{async_name} -> Page closed for {target_url}.")


# --- 主入口 ---
all_search_results_data = [] # Data collected from processed images

async def main():
    load_download_history(DOWNLOAD_HISTORY_FILE)
    load_url_history(HISTORY_IMG_URL_FILE) # Load the new URL-based history

    # Prompt for scroll attempts
    print("\n请选择滚动上限:")
    print("1. 30次 (默认)")
    print("2. 200次")
    print("3. 1000次")
    print("4. 自定义输入")
    
    choice = input("请输入您的选择 (1/2/3/4), 或直接回车选择默认 (30次): ").strip()
    
    selected_max_scroll_attempts = 30 # Default value
    if choice == '2':
        selected_max_scroll_attempts = 200
    elif choice == '3':
        selected_max_scroll_attempts = 1000
    elif choice == '4':
        while True:
            try:
                custom_input = input("请输入自定义滚动上限: ").strip()
                if custom_input:
                    selected_max_scroll_attempts = int(custom_input)
                    if selected_max_scroll_attempts <= 0:
                        print("输入无效，请重新输入一个正整数。")
                        continue
                else:
                    print("输入为空，将使用默认值30次。")
                break
            except ValueError:
                print("输入无效，请重新输入一个整数。")
    elif choice == '': # User pressed Enter
        print("已选择默认滚动上限: 30次。")
    else:
        print("无效选择，将使用默认滚动上限: 30次。")

    logger.info(f"Selected maximum scroll attempts: {selected_max_scroll_attempts}")

    # Initialize queues
    image_download_queue = asyncio.Queue()
    md5_analysis_queue = asyncio.Queue()

    # Start consumer tasks
    num_downloaders = 5 # Number of concurrent image downloaders
    num_md5_analyzers = 5 # Number of concurrent MD5 analyzers
    downloader_tasks = []
    analyzer_tasks = []

    for _ in range(num_downloaders):
        task = asyncio.create_task(image_downloader(image_download_queue, md5_analysis_queue))
        downloader_tasks.append(task)

    for _ in range(num_md5_analyzers):
        task = asyncio.create_task(md5_analyzer(md5_analysis_queue))
        analyzer_tasks.append(task)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            proxy={"server": PROXY} if PROXY else None,
            timeout=60000
        )
        context = await browser.new_context(
            viewport={'width': 2560, 'height': 1440}
        )
        target_urls = read_urls_from_file(KEYWORD_TARGET_FILE)
        if not target_urls:
            logger.error(f"No valid URLs found in {KEYWORD_TARGET_FILE}. Please add URLs to scrape.")
            # Signal workers to stop and then close browser
            for _ in range(num_downloaders):
                await image_download_queue.put(None)
            for _ in range(num_md5_analyzers):
                await md5_analysis_queue.put(None)
            await asyncio.gather(*downloader_tasks)
            await asyncio.gather(*analyzer_tasks)
            await browser.close()
            return
        
        # Start scraping tasks, passing the image_download_queue and selected_max_scroll_attempts
        scrape_tasks = [performCivitaiImageScrape(context, url, image_download_queue, selected_max_scroll_attempts) for url in target_urls]
        await asyncio.gather(*scrape_tasks)
        logger.info("All scraping tasks completed. Signaling download and analysis workers to stop.")

        # Signal consumer tasks to stop by putting None for each worker
        for _ in range(num_downloaders):
            await image_download_queue.put(None)
        for _ in range(num_md5_analyzers):
            await md5_analysis_queue.put(None)

        # Wait for all queued download and analysis tasks to be processed
        await asyncio.gather(*downloader_tasks)
        await asyncio.gather(*analyzer_tasks)

        await browser.close()
        logger.info("Browser closed. Script finished scraping and processing data.")

    # --- Excel 导出 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Civitai图片结果"
    headers = ["抓取时间", "搜索URL", "缩略图URL", "本地缩略图路径", "本地缩略图超链接", "原始图片详情页链接", "点赞数", "爱心数", "笑哭数", "伤心数", "打赏数", "关键词"]
    ws.append(headers)
    hyperlink_font = Font(color="0000FF", underline="single")
    for row_data in all_search_results_data:
        row = []
        for header in headers:
            if header == "本地缩略图超链接":
                row.append("点击打开缩略图")
            else:
                row.append(row_data.get(header, ""))
        ws.append(row)
        current_row_idx = ws.max_row
        search_url = row_data.get("搜索URL")
        if search_url:
            cell_search_url = ws.cell(row=current_row_idx, column=headers.index("搜索URL") + 1)
            cell_search_url.value = search_url
            cell_search_url.hyperlink = search_url
            cell_search_url.font = hyperlink_font
        thumbnail_url = row_data.get("缩略图URL")
        if thumbnail_url:
            cell_thumbnail_url = ws.cell(row=current_row_idx, column=headers.index("缩略图URL") + 1)
            cell_thumbnail_url.value = thumbnail_url
            cell_thumbnail_url.hyperlink = thumbnail_url
            cell_thumbnail_url.font = hyperlink_font
        local_image_hyperlink_url = row_data.get("本地缩略图超链接")
        if local_image_hyperlink_url:
            cell_local_image_hyperlink = ws.cell(row=current_row_idx, column=headers.index("本地缩略图超链接") + 1)
            cell_local_image_hyperlink.value = "点击打开缩略图"
            cell_local_image_hyperlink.hyperlink = local_image_hyperlink_url # 直接使用字符串URL
            cell_local_image_hyperlink.font = hyperlink_font
        original_page_link = row_data.get("原始图片详情页链接")
        if original_page_link:
            cell_original_page_link = ws.cell(row=current_row_idx, column=headers.index("原始图片详情页链接") + 1)
            cell_original_page_link.value = original_page_link
            cell_original_page_link.hyperlink = original_page_link
            cell_original_page_link.font = hyperlink_font
    wb.save(excel_filename)
    logger.info(f"Results saved to Excel: {excel_filename}")
    save_download_history(DOWNLOAD_HISTORY_FILE)
    save_url_history(HISTORY_IMG_URL_FILE) # Save the new URL-based history

if __name__ == '__main__':
    history_copy_filepath = None # Declare outside try block for finally access
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Script interrupted by user.")
    except Exception as e:
        logger.critical(f"An unhandled error occurred in main: {e}\n{traceback.format_exc()}")
    finally:
        for handler in logger.handlers[:]:
            try:
                handler.flush()
                handler.close()
                logger.removeHandler(handler)
            except Exception as e:
                print(f"Error closing log handler: {e}")
        try:
            if os.path.exists(log_filename):
                print(f"Attempting to open log file: {log_filename}")
                if os.name == 'nt':
                    os.startfile(log_filename)
                elif hasattr(os, "uname") and os.uname().sysname == 'Darwin':
                    subprocess.run(['open', log_filename])
                else:
                    subprocess.run(['xdg-open', log_filename])
            else:
                print(f"Log file not found: {log_filename}")
        except Exception as e:
            print(f"Error opening log file {log_filename}: {e}")
        try:
            if os.path.exists(excel_filename):
                print(f"Attempting to open Excel file: {excel_filename}")
                if os.name == 'nt':
                    os.startfile(excel_filename)
                elif hasattr(os, "uname") and os.uname().sysname == 'Darwin':
                    subprocess.run(['open', excel_filename])
                else:
                    subprocess.run(['xdg-open', excel_filename])
            else:
                print(f"Excel file not found: {excel_filename}")
        except Exception as e:
            print(f"Error opening Excel file {excel_filename}: {e}")
        
        # --- 新增：复制并打开带时间戳的历史记录文件 ---
        try:
            if os.path.exists(HISTORY_IMG_URL_FILE):
                # 确保 historyImgMD5 目录存在
                if not os.path.exists(HISTORY_MD5_DIR):
                    os.makedirs(HISTORY_MD5_DIR)
                
                # 构建带时间戳的文件名
                history_copy_filename = f"history_img_url_history_{timestamp}.xlsx"
                history_copy_filepath = os.path.join(HISTORY_MD5_DIR, history_copy_filename)
                
                # 复制文件
                shutil.copy2(HISTORY_IMG_URL_FILE, history_copy_filepath)
                logger.info(f"Copied URL history to {history_copy_filepath}")
                
                print(f"Attempting to open URL history Excel file copy: {history_copy_filepath}")
                if os.name == 'nt':
                    os.startfile(history_copy_filepath) # 使用文件关联打开
                elif hasattr(os, "uname") and os.uname().sysname == 'Darwin':
                    subprocess.run(['open', history_copy_filepath]) # 使用文件关联打开
                else:
                    subprocess.run(['xdg-open', history_copy_filepath]) # 使用文件关联打开
            else:
                print(f"Original URL history Excel file not found: {HISTORY_IMG_URL_FILE}. No copy made or opened.")
        except Exception as e:
            print(f"Error handling URL history Excel file copy and open: {e}")
            logger.error(f"Error handling URL history Excel file copy and open: {e}\n{traceback.format_exc()}")